#!/usr/bin/env python3
"""
fetch_data.py — Strategic Inputs Policy Monitor
Daily data fetch for the GTA dashboard. Writes:
  data/dashboard.json       — indicator counts consumed by index.html
  data/interventions.xlsx   — downloadable intervention records

Indicator methodology verified against GTA MCP on 2026-05-14.
"""

import json, os, sys, time
import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone

# ── Config ─────────────────────────────────────────────────────────────────────
API_KEY    = os.environ.get("GTA_API_KEY")
COUNTS_URL = "https://api.globaltradealert.org/api/v1/gta/data-counts/"
DATA_URL   = "https://api.globaltradealert.org/api/v2/gta/data/"
CUTOFF     = "2026-02-28"
END_OPEN   = "2099-12-31"
DELAY      = 0.5   # seconds between requests

# GTA evaluation IDs
HARMFUL_IDS      = {1, 2}   # Red, Amber
LIBERALISING_IDS = {3, 5}   # Green, Liberalising
MAST_L_ID        = 10       # Chapter L: Subsidies (excl. export subsidies)

# ── Products ───────────────────────────────────────────────────────────────────
PRODUCTS = [
    {
        "name": "Fuels",
        "hs_codes": [270900, 271012, 271019],
        "description": (
            "Direct exposure to crude supply disruptions, reduced refinery throughput, "
            "and constrained petroleum product exports through the Gulf/Hormuz corridor."
        ),
    },
    {
        "name": "Fertilizers",
        "hs_codes": [310210, 310221, 310310, 310520, 310530],
        "description": (
            "Strong dependence on natural gas-derived ammonia and refinery-derived sulphur "
            "used in urea, ammonium sulphate, phosphates, DAP/MAP, and NPK production."
        ),
    },
    {
        "name": "Sulphur",
        "hs_codes": [250300],
        "description": (
            "Sulphur is primarily recovered from oil refining and gas processing; "
            "reduced refinery operations directly constrain supply."
        ),
    },
    {
        "name": "Methanol",
        "hs_codes": [290511],
        "description": (
            "Produced mainly from natural gas feedstocks; Middle East production and "
            "export infrastructure highly exposed to Gulf disruptions."
        ),
    },
    {
        "name": "Graphite Feedstocks",
        "hs_codes": [271311, 271312, 380110],
        "description": (
            "Petroleum coke-based graphite feedstocks depend on refinery output "
            "and delayed coking capacity."
        ),
    },
    {
        "name": "Alumina",
        "hs_codes": [281820],
        "description": (
            "Energy-intensive refining process indirectly exposed to higher fuel and gas "
            "prices caused by regional energy disruptions."
        ),
    },
    {
        "name": "Helium",
        "hs_codes": [280429],
        "description": (
            "Significant global supply originates from Qatar and Gulf gas-processing "
            "infrastructure dependent on Hormuz shipping routes."
        ),
    },
    {
        "name": "Monoethylene Glycol (MEG)",
        "hs_codes": [290531],
        "description": (
            "Petrochemical derivative produced from ethylene; exposed to disruptions "
            "in naphtha and gas-based cracker feedstocks."
        ),
    },
    {
        "name": "Iron Ore",
        "hs_codes": [260111, 260112],
        "description": (
            "Primarily indirect exposure via higher bunker/freight costs and weaker "
            "steel-sector demand rather than refinery dependence."
        ),
    },
]

# ── Policy group definitions ────────────────────────────────────────────────────
# export_controls, import_barriers, export_subsidies, sanctions → matched by
# intervention_type_name (single string per record in the GTA API).
# domestic_subsidies → matched by mast_chapter_id == MAST_L_ID.
# other → type NOT in any named group AND mast_chapter_id != MAST_L_ID.

EXPORT_CONTROLS_TYPES = {
    "Export ban", "Export quota", "Export tax", "Export tariff quota",
    "Export licensing requirement", "Export price benchmark",
    "Local supply for exports", "Local supply requirement",
}
IMPORT_BARRIERS_TYPES = {
    "Import tariff", "Import quota", "Import ban", "Import tariff quota",
    "Import licensing requirement", "Import price benchmark",
    "Minimum import price", "Other import charges",
    "Internal taxation of imports", "Selective import channel restriction",
}
EXPORT_SUBSIDIES_TYPES = {
    "Export subsidy", "Trade finance",
    "Financial assistance in foreign market", "Other export incentive",
}
SANCTIONS_TYPES = {
    "Controls on commercial transactions and investment instruments",
}
ALL_NAMED_TYPES = (
    EXPORT_CONTROLS_TYPES | IMPORT_BARRIERS_TYPES |
    EXPORT_SUBSIDIES_TYPES | SANCTIONS_TYPES
)

# ── Helpers ────────────────────────────────────────────────────────────────────
def get_headers():
    return {
        "Content-Type": "application/json",
        "Authorization": f"APIKey {API_KEY}",
    }

def last_saturday(ref):
    """Return YYYY-MM-DD of the most recent Saturday on or before `ref`."""
    days_back = (ref.weekday() - 5) % 7
    return (ref - timedelta(days=days_back)).strftime("%Y-%m-%d")

def counts_request(hs_codes, date_end=None, count_by=None):
    """
    Call the data-counts endpoint with a 3-way count_by:
      ["gta_evaluation", "intervention_type", "mast_chapter"]
    Returns the results list.  Each row has:
      value, gta_evaluation_id, intervention_type_name, mast_chapter_id.
    """
    body = {
        "request_data": {
            "affected_products":  hs_codes,
            "announcement_period": [CUTOFF, date_end or END_OPEN],
            "count_by":           count_by or ["gta_evaluation", "intervention_type", "mast_chapter"],
            "count_variable":     "intervention_id",
        }
    }
    time.sleep(DELAY)
    r = requests.post(COUNTS_URL, headers=get_headers(), json=body, timeout=30)
    r.raise_for_status()
    return r.json().get("results", [])

def _get_mast_id(row):
    """Extract MAST chapter ID from a count_by row — handles multiple API response shapes."""
    # Flat integer (MCP response shape)
    v = row.get("mast_chapter_id")
    if isinstance(v, int):
        return v
    # Nested dict  e.g. {"id": 10, "name": "L: ..."}
    raw = row.get("mast_chapter") or v
    if isinstance(raw, dict):
        return raw.get("id") or raw.get("mast_chapter_id")
    if isinstance(raw, int):
        return raw
    return None

def _get_type_name(row):
    """Extract intervention type name — handles multiple API response shapes."""
    # Flat string (MCP and REST count_by response)
    v = row.get("intervention_type_name")
    if isinstance(v, str):
        return v
    raw = row.get("intervention_type")
    if isinstance(raw, str):
        return raw
    if isinstance(raw, dict):
        return raw.get("name", "")
    return ""

def _get_eval_id(row):
    """Extract GTA evaluation ID."""
    v = row.get("gta_evaluation_id")
    if isinstance(v, int):
        return v
    raw = row.get("gta_evaluation")
    if isinstance(raw, dict):
        return raw.get("id")
    return None

def compute_indicators(rows, total):
    """
    Derive all dashboard indicators from count_by rows.
    Returns dict with evaluation counts and policy group counts.

    Key invariant (verified on all 9 products):
      harmful + liberalising == total
    """
    harmful = liberalising = 0
    export_controls = import_barriers = domestic_subsidies = 0
    export_subsidies = sanctions = other = 0

    # Debug: print first row field names to catch API shape changes
    if rows:
        print(f"    [debug] count_by row keys: {list(rows[0].keys())}")
        print(f"    [debug] first row sample: {rows[0]}")

    for row in rows:
        v  = row.get("value", 0)
        ev = _get_eval_id(row)
        t  = _get_type_name(row)
        m  = _get_mast_id(row)

        # Evaluation
        if ev in HARMFUL_IDS:        harmful      += v
        elif ev in LIBERALISING_IDS: liberalising += v

        # Policy groups
        if t in EXPORT_CONTROLS_TYPES:  export_controls    += v
        if t in IMPORT_BARRIERS_TYPES:  import_barriers    += v
        if m == MAST_L_ID:              domestic_subsidies += v
        if t in EXPORT_SUBSIDIES_TYPES: export_subsidies   += v
        if t in SANCTIONS_TYPES:        sanctions          += v
        if t not in ALL_NAMED_TYPES and m != MAST_L_ID:
            other += v

    # Quality check
    qc = (harmful + liberalising) == total
    if not qc:
        print(f"    WARNING: QC FAIL — harmful({harmful}) + lib({liberalising}) = "
              f"{harmful+liberalising} ≠ total({total})")

    return {
        "evaluation":    {"harmful": harmful, "liberalising": liberalising},
        "policy_groups": {
            "export_controls":    export_controls,
            "import_barriers":    import_barriers,
            "domestic_subsidies": domestic_subsidies,
            "export_subsidies":   export_subsidies,
            "sanctions":          sanctions,
            "other":              other,
        },
    }

def fetch_wow(hs_codes, today):
    """Week-on-week change between two consecutive Saturday totals."""
    sat_this = last_saturday(today)
    sat_prev = last_saturday(today - timedelta(days=7))

    def total_at(date_end):
        rows = counts_request(hs_codes, date_end=date_end,
                              count_by=["gta_evaluation"])
        return sum(r.get("value", 0) for r in rows)

    n_this = total_at(sat_this)
    n_prev = total_at(sat_prev)
    return n_this - n_prev, sat_this, sat_prev

def fetch_implementing(hs_codes):
    """ISO alpha-3 → count dict for implementing jurisdictions."""
    rows = counts_request(hs_codes, count_by=["implementer"])
    result = {}
    for r in rows:
        iso = r.get("implementer_iso")
        cnt = r.get("value", 0)
        if iso and cnt:
            result[iso] = result.get(iso, 0) + cnt
    return result

def fetch_all_interventions(hs_codes):
    """
    Fetch every intervention record for the given HS codes from the v2 data
    endpoint.  Returns a list of raw record dicts.

    NOTE: keep_* flags are NOT set here.  Earlier testing showed that adding
    keep_affected / keep_affected_products caused the MCP search to return
    fewer records than the authoritative count.  Without these flags the v2
    endpoint returns the full count; affected_jurisdiction and affected_products
    may therefore be absent from some records — the script handles that
    gracefully by falling back to the intervention URL.
    """
    all_results = []
    offset, limit = 0, 1000
    first_batch = True

    while True:
        body = {
            "limit":  limit,
            "offset": offset,
            "request_data": {
                "affected_products":   hs_codes,
                "announcement_period": [CUTOFF, END_OPEN],
            },
        }
        time.sleep(DELAY)
        r = requests.post(DATA_URL, headers=get_headers(), json=body, timeout=60)
        r.raise_for_status()
        raw = r.json()

        results = raw if isinstance(raw, list) else raw.get("results", [])
        total   = len(raw) if isinstance(raw, list) else raw.get("count", 0)

        if first_batch and results:
            print(f"    API keys: {sorted(results[0].keys())}")
            first_batch = False

        all_results.extend(results)
        offset += limit
        if offset >= total or not results:
            break

    return all_results

# ── Field extraction ──────────────────────────────────────────────────────────
def safe_str(val):
    if val is None:                return ""
    if isinstance(val, str):       return val
    if isinstance(val, dict):      return val.get("name") or val.get("label") or str(val)
    if isinstance(val, list):      return "; ".join(safe_str(v) for v in val if v)
    return str(val)

def extract_row(rec):
    """Return an 8-element list matching XLSX_HEADERS."""
    iid   = rec.get("intervention_id", "")
    date  = rec.get("date_announced", "")
    title = rec.get("state_act_title") or rec.get("title") or rec.get("name") or ""
    url   = rec.get("intervention_url") or rec.get("url") or ""
    ev    = safe_str(rec.get("gta_evaluation", ""))

    # Implementing jurisdiction
    impl  = (
        safe_str(rec.get("implementing_jurisdictions"))
        or safe_str(rec.get("implementing_jurisdiction"))
        or safe_str(rec.get("implementer"))
        or ""
    )

    # Intervention type — singular string in the GTA v2 API
    itype = (
        safe_str(rec.get("intervention_type"))
        or safe_str(rec.get("intervention_types"))
        or ""
    )

    # Affected jurisdictions
    aff_j = (
        safe_str(rec.get("affected_jurisdictions"))
        or safe_str(rec.get("affected"))
        or ""
    )

    # Affected HS codes — only product_id (numeric), no descriptions
    prods_raw = rec.get("affected_products", [])
    if isinstance(prods_raw, list):
        hs_parts = []
        for p in prods_raw:
            if isinstance(p, dict):
                pid = p.get("product_id") or p.get("id") or ""
                hs_parts.append(str(pid) if pid else "")
            else:
                hs_parts.append(str(p))
        hs_codes_str = "; ".join(h for h in hs_parts if h)
    else:
        hs_codes_str = safe_str(prods_raw)

    return [iid, date, title, impl, ev, itype, aff_j, hs_codes_str, url]

# ── Excel generation ───────────────────────────────────────────────────────────
XLSX_HEADERS = [
    "Intervention ID", "Announcement Date", "Title",
    "Implementing Jurisdiction", "GTA Evaluation", "GTA Intervention Type",
    "Affected Jurisdiction", "Affected HS Codes", "Intervention URL",
]
COL_W = [16, 16, 65, 55, 16, 40, 55, 35, 60]
NAVY  = "0E334D"
WHITE = "FFFFFF"
LIGHT = "E8EEF5"
RED_C = "C03838"
GREEN_C = "28A068"
AMBER_C = "D4941A"

def style_header(ws, row_num=1):
    for c in range(1, len(XLSX_HEADERS) + 1):
        cell = ws.cell(row_num, c)
        cell.font      = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 18

def set_widths(ws):
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_sheet(wb, name, rows, product_label=None, download_date=None):
    ws = wb.create_sheet(name[:31])
    ws.freeze_panes = "A2"

    if product_label:
        ws.append([f"Product category: {product_label}"])
        ws["A1"].font = Font(bold=True, color=NAVY, name="Calibri", size=10)
        ws.append(XLSX_HEADERS)
        style_header(ws, row_num=2)
        data_start = 3
    else:
        ws.append(XLSX_HEADERS)
        style_header(ws, row_num=1)
        data_start = 2

    EC = {"Red": RED_C, "Amber": AMBER_C, "Green": GREEN_C}

    for i, row in enumerate(rows, data_start):
        ws.append(row)
        bg = LIGHT if (i - data_start) % 2 == 1 else None
        for c in range(1, len(XLSX_HEADERS) + 1):
            cell = ws.cell(i, c)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=False)
            if bg:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            if c == 5:  # eval colour
                cell.font = Font(bold=True, color=EC.get(row[4], "000000"),
                                 name="Calibri", size=10)
            elif c == 9:  # URL hyperlink
                cell.hyperlink = row[8] if row[8] else None
                cell.font = Font(color="1874CD", underline="single",
                                 name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)

    set_widths(ws)

    # Download date at bottom
    if download_date:
        blank_row = data_start + len(rows)
        ws.append([])
        date_row = blank_row + 1
        cell = ws.cell(date_row, 1)
        cell.value = f"Downloaded: {download_date}"
        cell.font  = Font(italic=True, color="636878", name="Calibri", size=9)
        ws.merge_cells(
            start_row=date_row, start_column=1,
            end_row=date_row, end_column=len(XLSX_HEADERS)
        )

    return ws

# ── Per-product computation ────────────────────────────────────────────────────
def compute_product(product, today):
    codes = product["hs_codes"]
    name  = product["name"]

    # ── Counts via 3-way count_by (authoritative for indicators) ──────────────
    print(f"    counts ...", end=" ", flush=True)
    rows_3way = counts_request(codes)
    total = sum(r.get("value", 0) for r in rows_3way)
    print(f"total={total}", end="  ", flush=True)

    indicators = compute_indicators(rows_3way, total)
    print(f"H={indicators['evaluation']['harmful']} "
          f"L={indicators['evaluation']['liberalising']} | "
          f"groups: " +
          " ".join(f"{k[:3]}={v}" for k, v in indicators["policy_groups"].items()))

    # ── Week-on-week ───────────────────────────────────────────────────────────
    print(f"    wow ...", end=" ", flush=True)
    wow, sat_this, sat_prev = fetch_wow(codes, today)
    print(f"{wow:+d}  (ref: {sat_prev} → {sat_this})")

    # ── Implementing jurisdictions ─────────────────────────────────────────────
    print(f"    implementing ...", end=" ", flush=True)
    implementing = fetch_implementing(codes)
    print(f"{len(implementing)} countries")

    # ── Individual records for xlsx ────────────────────────────────────────────
    print(f"    records ...", end=" ", flush=True)
    records = fetch_all_interventions(codes)
    print(f"{len(records)} records")

    # Top 5 most recent interventions for dashboard slide-out
    sorted_recs = sorted(records, key=lambda r: r.get("date_announced",""), reverse=True)
    latest = []
    for rec in sorted_recs[:5]:
        latest.append({
            "date":              rec.get("date_announced",""),
            "title":             rec.get("state_act_title") or rec.get("title",""),
            "intervention_type": rec.get("intervention_type",""),
            "gta_evaluation":    rec.get("gta_evaluation",""),
            "url":               rec.get("intervention_url") or rec.get("url",""),
        })

    entry = {
        "name":                  product["name"],
        "hs_codes":              codes,
        "description":           product["description"],
        "total_interventions":   total,
        "wow_change":            wow,
        "reference_saturdays":   {"current": sat_this, "previous": sat_prev},
        "evaluation":            indicators["evaluation"],
        "policy_groups":         indicators["policy_groups"],
        "implementing":          implementing,
        "latest_interventions":  latest,
    }

    return entry, records

# ── Excel builder ──────────────────────────────────────────────────────────────
def build_excel(interventions_by_product, download_date):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # All Products — deduplicated
    seen, all_rows = set(), []
    for product_name, records in interventions_by_product.items():
        for rec in records:
            iid = rec.get("intervention_id")
            if iid not in seen:
                seen.add(iid)
                all_rows.append(extract_row(rec))
    add_sheet(wb, "All Products", all_rows, download_date=download_date)
    print(f"  All Products: {len(all_rows)} unique interventions")

    # Per-product tabs
    for product_name, records in interventions_by_product.items():
        rows = [extract_row(r) for r in records]
        add_sheet(wb, product_name[:31], rows,
                  product_label=product_name, download_date=download_date)
        print(f"  {product_name}: {len(rows)} interventions")

    os.makedirs("data", exist_ok=True)
    wb.save("data/interventions.xlsx")
    print("Saved → data/interventions.xlsx")

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if not API_KEY:
        print("ERROR: GTA_API_KEY is not set.", file=sys.stderr)
        sys.exit(1)

    today = datetime.now(timezone.utc)
    run_ts = today.strftime("%Y-%m-%d %H:%M UTC")
    download_date = today.strftime("%d %B %Y")

    print("=== Strategic Inputs Policy Monitor — Data Fetch ===")
    print(f"Cutoff : {CUTOFF}")
    print(f"Run    : {run_ts}")
    print()

    output = {
        "last_updated": run_ts,
        "cutoff_date":  CUTOFF,
        "overview": (
            "This dashboard tracks trade policy interventions affecting strategic input "
            "commodities with direct or indirect exposure to the 2026 Iran–Hormuz crisis. "
            "All indicators count interventions announced since 28 February 2026, as "
            "recorded in the Global Trade Alert database. Data is refreshed daily."
        ),
        "products": [],
    }

    interventions_by_product = {}

    for i, product in enumerate(PRODUCTS, 1):
        print(f"[{i}/{len(PRODUCTS)}] {product['name']}")
        entry, records = compute_product(product, today)
        output["products"].append(entry)
        interventions_by_product[product["name"]] = records
        print()

    # Write dashboard.json first — always saved even if xlsx fails
    os.makedirs("data", exist_ok=True)
    with open("data/dashboard.json", "w") as f:
        json.dump(output, f, indent=2)
    print("Saved → data/dashboard.json\n")

    # Write interventions xlsx
    print("Building Excel file...")
    build_excel(interventions_by_product, download_date)


if __name__ == "__main__":
    main()
